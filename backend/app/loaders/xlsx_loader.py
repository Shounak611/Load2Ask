from pathlib import Path
from typing import Union, Dict, Any
from app.loaders.base import BaseLoader
from app.models.internal import Document
from app.core.errors import InvalidFileError
from app.core.logging import logger

try:
    import openpyxl
except ImportError:
    openpyxl = None


class XLSXLoader(BaseLoader):
    """Excel (XLSX) loader extracting sheet names, headers, and structured rows."""

    def load(self, source: Union[str, Path], metadata: Dict[str, Any] = None) -> Document:
        file_path = Path(source)
        if not file_path.exists():
            raise InvalidFileError(f"XLSX file not found: {source}")

        if openpyxl is None:
            raise InvalidFileError("openpyxl package is not installed.")

        try:
            wb = openpyxl.load_workbook(str(file_path), data_only=True)
            sheet_contents = []
            all_columns = []
            sheet_names = wb.sheetnames

            for sheet_name in sheet_names:
                sheet = wb[sheet_name]
                rows = list(sheet.iter_rows(values_only=True))
                if not rows:
                    continue

                header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
                all_columns.extend(header)

                sheet_lines = [f"[Sheet: {sheet_name}]", f"Columns: {', '.join(header)}"]
                for r_idx, row in enumerate(rows[1:], start=1):
                    row_parts = []
                    for c_idx, cell_val in enumerate(row):
                        if cell_val is not None and str(cell_val).strip():
                            col_name = header[c_idx] if c_idx < len(header) else f"col_{c_idx+1}"
                            row_parts.append(f"{col_name}: {str(cell_val).strip()}")
                    if row_parts:
                        sheet_lines.append(f"[Row {r_idx}] " + " | ".join(row_parts))

                sheet_contents.append("\n".join(sheet_lines))

            full_content = "\n\n".join(sheet_contents)

            meta = {
                "source_type": "xlsx",
                "file_name": file_path.name,
                "sheets": sheet_names,
                "columns": list(set(all_columns)),
                "file_size": file_path.stat().st_size,
                **(metadata or {})
            }

            return Document(
                source_type="xlsx",
                source_name=file_path.name,
                source_uri=str(file_path.absolute()),
                content=full_content,
                metadata=meta
            )
        except Exception as e:
            logger.error(f"Failed to load XLSX file {source}: {e}")
            raise InvalidFileError(f"Failed to read XLSX file {source}: {str(e)}")
